#!/usr/bin/env python3
"""
CLEARLINE CLAIMS BATCH PROCESSOR
==================================
Reads a hospital claims Excel/CSV file, groups rows into PA batches
by enrollee + encounter date, resolves procedure/diagnosis descriptions
to codes, submits each batch to the vetting API, and outputs a detailed
savings report.

Batching logic (per Clearline PA spec):
  - Rows with PA Number != 0  → anchor rows (one batch per PA number)
  - Rows with PA Number == 0  → attach to anchor with same enrollee + date
  - If ALL rows are PA=0      → group by enrollee + encounter date

Usage (CLI):
    python -m apis.vetting.claims_processor claims.xlsx \\
        --provider-id 118 \\
        --hospital "R-Jolad Hospital" \\
        --api http://localhost:8000 \\
        --output report.xlsx

Author: Clearline AI
"""

import os
import re
import sys
import json
import argparse
import logging
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple

import pandas as pd
import requests

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")

# ─────────────────────────────────────────────────────────────────────────────
# KNOWN-GOOD NAME → CODE MAPPINGS
# Used when DB lookup fails or for speed. Add to these as you learn more codes.
# ─────────────────────────────────────────────────────────────────────────────

PROC_NAME_MAP: Dict[str, str] = {
    # Only REAL procedure codes that exist in the Clearline DB.
    # Everything else falls back to using the description as the code —
    # that way the API receives the human-readable name and all
    # name-based checks (vitamin padding, IV fluid, shotgun labs, etc.) fire correctly.

    # GP / General Practice
    "GP CONSULTATION":                          "CONS021",
    "GENERAL PRACTITIONER CONSULTATION":        "CONS021",
    "GENERAL PRACTICE CONSULTATION":            "CONS021",
    "GENERAL PRACTITIONER":                     "CONS021",
}

DIAG_NAME_MAP: Dict[str, str] = {
    # Malaria
    "UNSPECIFIED MALARIA":                              "B54",
    "MALARIA, UNSPECIFIED":                             "B54",
    "MALARIA UNSPECIFIED":                              "B54",
    "PLASMODIUM FALCIPARUM MALARIA, UNSPECIFIED":       "B50.9",
    "PLASMODIUM VIVAX MALARIA":                         "B51.9",
    # URTI
    "ACUTE UPPER RESPIRATORY INFECTION, UNSPECIFIED":   "J06.9",
    "ACUTE UPPER RESPIRATORY INFECTION":                "J06.9",
    "UPPER RESPIRATORY TRACT INFECTION":                "J06.9",
    "UPPER RESPIRATORY INFECTION":                      "J06.9",
    "URTI":                                             "J06.9",
    # Skin
    "DIAPER DERMATITIS":                                "L22",
    "NAPPY RASH":                                       "L22",
    "DIAPER RASH":                                      "L22",
    # GI
    "GASTROENTERITIS":                                  "A09",
    "ACUTE GASTROENTERITIS":                            "A09",
    "DIARRHOEA":                                        "A09",
    # Hypertension
    "ESSENTIAL HYPERTENSION":                           "I10",
    "HYPERTENSION":                                     "I10",
    # DM
    "TYPE 2 DIABETES MELLITUS":                         "E11.9",
    "DIABETES MELLITUS, TYPE 2":                        "E11.9",
    # Typhoid
    "TYPHOID FEVER":                                    "A01.0",
    # UTI
    "URINARY TRACT INFECTION":                          "N39.0",
    "UTI":                                              "N39.0",
    # Anaemia
    "ANAEMIA, UNSPECIFIED":                             "D64.9",
    "ANAEMIA":                                          "D64.9",
    "ANEMIA":                                           "D64.9",
    # Pharyngitis / tonsillitis
    "ACUTE PHARYNGITIS":                                "J02.9",
    "ACUTE TONSILLITIS":                                "J03.9",
    # Fever
    "FEVER":                                            "R50.9",
    "PYREXIA OF UNKNOWN ORIGIN":                        "R50.9",
    "FEVER, UNSPECIFIED":                               "R50.9",
    # Malnutrition
    "PROTEIN-ENERGY MALNUTRITION":                      "E46",
    "MALNUTRITION":                                     "E46",
}


# ─────────────────────────────────────────────────────────────────────────────
# DB-ASSISTED CODE RESOLVER
# ─────────────────────────────────────────────────────────────────────────────

class CodeResolver:
    """
    Resolves procedure / diagnosis free-text descriptions to internal codes.
    Uses MotherDuck for fuzzy lookup; falls back to built-in maps then raw description.
    """

    def __init__(self, db_path: Optional[str] = None):
        self._conn = None
        self._db_path = db_path
        self._proc_cache: Dict[str, str] = {}
        self._diag_cache: Dict[str, str] = {}

    def _get_conn(self):
        if self._conn is not None:
            return self._conn
        if not self._db_path:
            return None
        try:
            import duckdb
            self._conn = duckdb.connect(self._db_path, read_only=True)
            logger.info("CodeResolver: MotherDuck connected")
        except Exception as e:
            logger.warning(f"CodeResolver: DB connection failed ({e}) — using name maps only")
        return self._conn

    @staticmethod
    def _norm(s: str) -> str:
        return re.sub(r'\s+', ' ', str(s).upper().strip())

    def resolve_procedure(self, description: str) -> Tuple[str, str]:
        """
        Returns (code, resolved_name).
        Priority: cache → built-in map → DB LIKE → raw description as code.
        """
        key = self._norm(description)

        if key in self._proc_cache:
            return self._proc_cache[key], description

        # Built-in map (partial match)
        for pattern, code in PROC_NAME_MAP.items():
            if pattern in key:
                self._proc_cache[key] = code
                return code, description

        # DB fuzzy lookup
        conn = self._get_conn()
        if conn:
            try:
                # Try exact LIKE match first
                row = conn.execute("""
                    SELECT TRIM(procedurecode) as code, TRIM(proceduredesc) as name
                    FROM "AI DRIVEN DATA"."PROCEDURE DATA"
                    WHERE UPPER(TRIM(proceduredesc)) LIKE ?
                    LIMIT 1
                """, [f"%{key[:30]}%"]).fetchone()
                if row and row[0]:
                    code = str(row[0]).strip().upper()
                    self._proc_cache[key] = code
                    return code, str(row[1]).strip()
            except Exception as e:
                logger.debug(f"Procedure DB lookup error: {e}")

        # Fall back: use description as code (system resolves name from it)
        self._proc_cache[key] = key
        return key, description

    def resolve_diagnosis(self, description: str) -> Tuple[str, str]:
        """
        Returns (icd10_code, resolved_name).
        Priority: cache → built-in map → DB LIKE → raw description as code.
        """
        if not description or str(description).strip() in ("", "nan"):
            return "", ""

        key = self._norm(description)

        if key in self._diag_cache:
            return self._diag_cache[key], description

        # Built-in map (exact key match first, then partial)
        if key in DIAG_NAME_MAP:
            code = DIAG_NAME_MAP[key]
            self._diag_cache[key] = code
            return code, description

        for pattern, code in DIAG_NAME_MAP.items():
            if pattern in key or key in pattern:
                self._diag_cache[key] = code
                return code, description

        # DB fuzzy lookup
        conn = self._get_conn()
        if conn:
            try:
                row = conn.execute("""
                    SELECT TRIM(diagnosiscode) as code, TRIM(diagnosisdesc) as name
                    FROM "AI DRIVEN DATA"."DIAGNOSIS"
                    WHERE UPPER(TRIM(diagnosisdesc)) LIKE ?
                    LIMIT 1
                """, [f"%{key[:30]}%"]).fetchone()
                if row and row[0]:
                    code = str(row[0]).strip().upper()
                    self._diag_cache[key] = code
                    return code, str(row[1]).strip()
            except Exception as e:
                logger.debug(f"Diagnosis DB lookup error: {e}")

        # Fall back: use description as code
        self._diag_cache[key] = key
        return key, description


# ─────────────────────────────────────────────────────────────────────────────
# CLAIMS FILE PARSER
# ─────────────────────────────────────────────────────────────────────────────

COLUMN_ALIASES = {
    "enrollee no":        "enrollee_id",
    "enrollee number":    "enrollee_id",
    "enrollee id":        "enrollee_id",
    "insured id":         "enrollee_id",
    "pa number":          "pa_number",
    "pa no":              "pa_number",
    "pa_number":          "pa_number",
    "encounter date":     "encounter_date",
    "date":               "encounter_date",
    "service date":       "encounter_date",
    "amount charged":     "amount",
    "amount":             "amount",
    "charge":             "amount",
    "units":              "quantity",
    "quantity":           "quantity",
    "qty":                "quantity",
    "procedure description": "procedure_description",
    "procedure":          "procedure_description",
    "service":            "procedure_description",
    "description":        "procedure_description",
    "diagnosis":          "diagnosis",
    "primary diagnosis":  "diagnosis",
    "diagnosis description": "diagnosis",
    "additional diagnosis": "additional_diagnosis",
    "provider name":      "provider_name",
    "hospital":           "provider_name",
    "s/no":               "sno",
    "s/n":                "sno",
    "no":                 "sno",
}


def load_claims(path: str) -> pd.DataFrame:
    """Load and normalise a claims Excel or CSV file."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(path)
    elif ext == ".csv":
        df = pd.read_csv(path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

    # Normalise column names
    df.columns = [c.strip().lower() for c in df.columns]
    rename = {}
    for col in df.columns:
        if col in COLUMN_ALIASES:
            rename[col] = COLUMN_ALIASES[col]
    df.rename(columns=rename, inplace=True)

    required = {"enrollee_id", "encounter_date", "procedure_description", "diagnosis"}
    missing  = required - set(df.columns)
    if missing:
        raise ValueError(f"Claims file missing required columns: {missing}. Found: {list(df.columns)}")

    # Coerce types
    df["encounter_date"] = pd.to_datetime(df["encounter_date"], errors="coerce")
    df["pa_number"]      = pd.to_numeric(df.get("pa_number", 0), errors="coerce").fillna(0).astype(int)
    df["quantity"]       = pd.to_numeric(df.get("quantity", 1), errors="coerce").fillna(1).astype(int)
    df["amount"]         = (
        df["amount"].astype(str)
        .str.replace(r"[₦,\s]", "", regex=True)
        .pipe(pd.to_numeric, errors="coerce")
        .fillna(0.0)
        if "amount" in df.columns
        else pd.Series([0.0] * len(df))
    )
    df["diagnosis"]      = df["diagnosis"].fillna("").astype(str).str.strip()
    df["additional_diagnosis"] = (
        df["additional_diagnosis"].fillna("").astype(str).str.strip()
        if "additional_diagnosis" in df.columns
        else ""
    )
    df["provider_name"]  = (
        df["provider_name"].fillna("").astype(str).str.strip()
        if "provider_name" in df.columns
        else ""
    )

    # ── Enrollee ID resolution ──────────────────────────────────────────────
    # Some hospital files omit the enrollee ID column entirely.
    # When enrollee_id is blank, look up from PA DATA using the PA number.
    df["enrollee_id"] = df["enrollee_id"].fillna("").astype(str).str.strip()
    missing_mask = (df["enrollee_id"] == "") | (df["enrollee_id"].str.lower() == "nan")
    if missing_mask.any():
        logger.info(f"  {missing_mask.sum()} row(s) missing enrollee_id — resolving from PA DATA...")
        try:
            import duckdb as _ddb, os as _os
            _db_path = _os.getenv("DUCKDB_PATH", "ai_driven_data.duckdb")
            _conn = _ddb.connect(_db_path, read_only=True)
            pa_to_iid: dict = {}
            non_zero_pa = df.loc[missing_mask & (df["pa_number"] != 0), "pa_number"].unique()
            if len(non_zero_pa):
                placeholders = ",".join(f"'{int(p)}'" for p in non_zero_pa)
                rows = _conn.execute(f'''
                    SELECT DISTINCT panumber, IID
                    FROM "AI DRIVEN DATA"."PA DATA"
                    WHERE panumber IN ({placeholders})
                ''').fetchall()
                pa_to_iid = {str(r[0]): r[1] for r in rows}
                logger.info(f"  Resolved {len(pa_to_iid)} PA → enrollee mappings from PA DATA")

            def _resolve_eid(row):
                if not missing_mask.loc[row.name]:
                    return row["enrollee_id"]
                pa = str(int(row["pa_number"])) if row["pa_number"] != 0 else "0"
                return pa_to_iid.get(pa, "")

            df["enrollee_id"] = df.apply(_resolve_eid, axis=1)
            _conn.close()
        except Exception as e:
            logger.warning(f"  PA DATA enrollee lookup failed: {e}")

    df.dropna(subset=["encounter_date", "procedure_description"], inplace=True)
    df = df[df["enrollee_id"] != ""].copy()
    df["encounter_date_str"] = df["encounter_date"].dt.strftime("%Y-%m-%d")
    logger.info(f"Loaded {len(df)} claim rows from {os.path.basename(path)}")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# BATCH BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_batches(df: pd.DataFrame) -> List[Dict]:
    """
    Groups claim rows into vetting batches.

    Logic:
      • Rows where pa_number != 0 are anchors — one batch per anchor.
      • Rows where pa_number == 0 attach to an anchor with the same
        enrollee_id + encounter_date.
      • If no anchors exist in the file, every unique (enrollee_id, date)
        pair forms its own batch.

    Returns list of dicts:
      { enrollee_id, encounter_date, pa_number, provider_name, rows: [DataFrame rows] }
    """
    batches = []

    anchors   = df[df["pa_number"] != 0].copy()
    free_rows = df[df["pa_number"] == 0].copy()

    # Track which PA=0 rows have been claimed by an anchor batch
    claimed_indices: set = set()

    # ── Anchor batches (PA != 0) ──────────────────────────────────────────
    for pa_num, pa_group in anchors.groupby("pa_number", sort=False):
        eid  = str(pa_group["enrollee_id"].iloc[0]).strip()
        edt  = pa_group["encounter_date_str"].iloc[0]
        pano = int(pa_num)

        # PA=0 rows for the same enrollee + date attach to this anchor
        sibling_mask = (
            (free_rows["enrollee_id"].astype(str).str.strip() == eid) &
            (free_rows["encounter_date_str"] == edt)
        )
        siblings = free_rows[sibling_mask]
        claimed_indices.update(siblings.index.tolist())

        batch_rows = pd.concat([pa_group, siblings]).reset_index(drop=True)
        batches.append({
            "enrollee_id":    eid,
            "encounter_date": edt,
            "pa_number":      pano,
            "provider_name":  str(pa_group["provider_name"].iloc[0]).strip()
                              if "provider_name" in pa_group.columns else "",
            "rows":           batch_rows,
        })

    # ── Orphan PA=0 rows (no anchor on that enrollee+date) ───────────────
    # These are real visits without a PA number — group by enrollee + date
    orphans = free_rows[~free_rows.index.isin(claimed_indices)]
    if not orphans.empty:
        grouped = orphans.groupby(
            [orphans["enrollee_id"].astype(str).str.strip(), "encounter_date_str"],
            sort=False
        )
        for (eid, edt), group in grouped:
            pname = group["provider_name"].iloc[0] if "provider_name" in group.columns else ""
            batches.append({
                "enrollee_id":    eid,
                "encounter_date": edt,
                "pa_number":      0,
                "provider_name":  str(pname).strip(),
                "rows":           group.reset_index(drop=True),
            })

    logger.info(f"Built {len(batches)} batch(es) from claims file")
    return batches


# ─────────────────────────────────────────────────────────────────────────────
# API SUBMITTER
# ─────────────────────────────────────────────────────────────────────────────

def submit_batch(
    batch: Dict,
    resolver: CodeResolver,
    api_url: str,
    provider_id: Optional[str],
    hospital_name: Optional[str],
    timeout: int = 300,
) -> Dict:
    """
    Resolve codes and POST one batch to /api/v1/validate/bulk.
    Returns the API response dict with metadata attached.
    """
    procedures = []
    for _, row in batch["rows"].iterrows():
        proc_desc  = str(row["procedure_description"]).strip()
        diag_desc  = str(row["diagnosis"]).strip()
        addl_desc  = str(row.get("additional_diagnosis", "")).strip()

        proc_code, proc_name = resolver.resolve_procedure(proc_desc)
        diag_code, diag_name = resolver.resolve_diagnosis(diag_desc)

        # Use additional diagnosis if primary resolves to empty
        if not diag_code and addl_desc:
            diag_code, diag_name = resolver.resolve_diagnosis(addl_desc)

        qty    = int(row.get("quantity", 1)) or 1
        amount = float(row.get("amount", 0)) or None
        price  = round(amount / qty, 4) if (amount and qty) else None

        procedures.append({
            "procedure_code": proc_code or proc_desc,
            "diagnosis_code": diag_code or diag_desc,
            "price":          price,
            "quantity":       qty,
            "notes":          proc_desc,  # keep original name as note
        })

    # Pass the PA number so the frequency check can exclude the pre-auth record
    # for this exact visit (same PA number = same encounter, not a prior visit).
    pa_num = batch.get("pa_number", 0)
    payload = {
        "enrollee_id":    batch["enrollee_id"],
        "encounter_date": batch["encounter_date"],
        "hospital_name":  hospital_name or batch.get("provider_name") or "Unknown",
        "provider_id":    provider_id or None,
        "pa_number":      str(pa_num) if pa_num else None,
        "encounter_type": batch.get("encounter_type", "OUTPATIENT"),
        "procedures":     procedures,
    }

    logger.info(
        f"  → Submitting batch: {batch['enrollee_id']} | {batch['encounter_date']} "
        f"| {len(procedures)} procedure(s)"
    )

    try:
        resp = requests.post(
            f"{api_url.rstrip('/')}/api/v1/validate/bulk",
            json=payload,
            timeout=timeout,
        )
        resp.raise_for_status()
        result = resp.json()
        result["_batch_meta"] = {
            "enrollee_id":    batch["enrollee_id"],
            "encounter_date": batch["encounter_date"],
            "pa_number":      batch["pa_number"],
            "provider_name":  hospital_name or batch.get("provider_name", ""),
            "row_count":      len(procedures),
            "original_rows":  batch["rows"].to_dict("records"),
        }
        return result
    except requests.exceptions.ConnectionError:
        logger.error(f"  ✗ Cannot reach API at {api_url} — is the server running?")
        return _error_result(batch, f"API unreachable at {api_url}")
    except requests.exceptions.HTTPError as e:
        logger.error(f"  ✗ API error {e.response.status_code}: {e.response.text[:200]}")
        return _error_result(batch, f"API error: {e.response.status_code}")
    except Exception as e:
        logger.error(f"  ✗ Unexpected error: {e}")
        return _error_result(batch, str(e))


def _error_result(batch: Dict, error_msg: str) -> Dict:
    return {
        "overall_status": "ERROR",
        "overall_decision": "ERROR",
        "total_approved_amount": 0.0,
        "line_items": [],
        "error": error_msg,
        "_batch_meta": {
            "enrollee_id":   batch["enrollee_id"],
            "encounter_date": batch["encounter_date"],
            "pa_number":     batch["pa_number"],
            "provider_name": batch.get("provider_name", ""),
            "row_count":     len(batch["rows"]),
            "original_rows": batch["rows"].to_dict("records"),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# REPORT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def generate_report(results: List[Dict], output_path: Optional[str] = None) -> pd.DataFrame:
    """
    Build a summary DataFrame and optionally write a colour-coded Excel report.
    Also prints a console summary.
    """
    # ── Build flat line-level detail table ───────────────────────────────────
    detail_rows = []
    for result in results:
        meta        = result.get("_batch_meta", {})
        enrollee    = meta.get("enrollee_id", "")
        enc_date    = meta.get("encounter_date", "")
        pa_number   = meta.get("pa_number", 0)
        provider    = meta.get("provider_name", "")
        orig_rows   = meta.get("original_rows", [])

        # Map procedure_code back to original description for readability
        code_to_desc = {
            str(r.get("procedure_description", "")).strip().upper()[:20]: r.get("procedure_description", "")
            for r in orig_rows
        }
        orig_amounts = {
            str(r.get("procedure_description", "")).strip().upper(): float(str(r.get("amount", 0)).replace("₦","").replace(",","") or 0)
            for r in orig_rows
        }

        line_items = result.get("line_items", [])
        if not line_items and result.get("overall_status") == "ERROR":
            # Error batch — add one row
            detail_rows.append({
                "Provider":         provider,
                "Enrollee ID":      enrollee,
                "PA Number":        pa_number,
                "Encounter Date":   enc_date,
                "Procedure Code":   "—",
                "Procedure Name":   "BATCH ERROR",
                "Diagnosis Code":   "—",
                "Diagnosis Name":   "—",
                "Stated Amount (₦)": 0.0,
                "Approved Amount (₦)": 0.0,
                "Status":           "ERROR",
                "Stage":            "—",
                "Denial Reason":    result.get("error", "Unknown"),
            })
            continue

        for item in line_items:
            pname  = item.get("procedure_name") or item.get("procedure_code", "")
            stated = item.get("stated_price", 0) or 0
            qty    = item.get("stated_quantity", 1) or 1
            stated_total = stated * qty

            approved_total = item.get("total_amount") or 0.0
            if item.get("status") == "AUTO_DENIED":
                approved_total = 0.0

            # Try to get original amount from claims file
            notes = ""
            if item.get("rules"):
                denial = next((r for r in item["rules"] if not r.get("passed")), None)
                if denial:
                    notes = denial.get("reasoning", "")[:120]

            detail_rows.append({
                "Provider":           provider,
                "Enrollee ID":        enrollee,
                "PA Number":          pa_number,
                "Encounter Date":     enc_date,
                "Procedure Code":     item.get("procedure_code", ""),
                "Procedure Name":     pname,
                "Diagnosis Code":     item.get("diagnosis_code", ""),
                "Diagnosis Name":     item.get("diagnosis_name", ""),
                "Stated Amount (₦)":  round(stated_total, 2),
                "Approved Amount (₦)": round(approved_total, 2),
                "Status":             item.get("status", ""),
                "Stage":              item.get("pipeline_stage", ""),
                "Denial Reason":      item.get("drop_reason", "") or notes,
            })

    detail_df = pd.DataFrame(detail_rows)

    # ── Build batch-level summary ─────────────────────────────────────────────
    summary_rows = []
    total_claimed  = 0.0
    total_approved = 0.0

    for result in results:
        meta      = result.get("_batch_meta", {})
        enrollee  = meta.get("enrollee_id", "")
        enc_date  = meta.get("encounter_date", "")
        provider  = meta.get("provider_name", "")
        pa_number = meta.get("pa_number", 0)

        line_items   = result.get("line_items", [])
        n_approved   = sum(1 for l in line_items if l.get("status") == "AUTO_APPROVED")
        n_denied     = sum(1 for l in line_items if l.get("status") == "AUTO_DENIED")

        # Stated total from original rows
        orig = meta.get("original_rows", [])
        stated_total = sum(
            float(str(r.get("amount", 0)).replace("₦","").replace(",","") or 0)
            for r in orig
        )
        approved_total = result.get("total_approved_amount", 0.0) or 0.0
        saved          = stated_total - approved_total

        total_claimed  += stated_total
        total_approved += approved_total

        # First denial reason
        first_denial = ""
        for item in line_items:
            if item.get("status") == "AUTO_DENIED":
                first_denial = item.get("drop_reason") or item.get("reasoning", "")[:80]
                break

        summary_rows.append({
            "Provider":              provider,
            "Enrollee ID":           enrollee,
            "PA Number":             pa_number,
            "Encounter Date":        enc_date,
            "Lines":                 len(line_items),
            "Approved Lines":        n_approved,
            "Denied Lines":          n_denied,
            "Stated Total (₦)":      round(stated_total, 2),
            "Approved Total (₦)":    round(approved_total, 2),
            "Savings (₦)":           round(saved, 2),
            "Overall Status":        result.get("overall_status", ""),
            "Primary Denial Reason": first_denial,
        })

    summary_df = pd.DataFrame(summary_rows)

    # ── Console output ────────────────────────────────────────────────────────
    total_saved   = total_claimed - total_approved
    savings_pct   = (total_saved / total_claimed * 100) if total_claimed else 0

    print("\n" + "━"*66)
    print("  CLEARLINE CLAIMS VETTING REPORT")
    print("━"*66)
    print(f"  Batches processed  : {len(results)}")
    print(f"  Total lines        : {len(detail_df)}")
    print(f"  Total claimed      : ₦{total_claimed:>12,.2f}")
    print(f"  Total approved     : ₦{total_approved:>12,.2f}")
    print(f"  Total SAVINGS      : ₦{total_saved:>12,.2f}  ({savings_pct:.1f}%)")
    print("━"*66)
    print("\nBatch Summary:")
    for r in summary_rows:
        status_icon = "✅" if r["Overall Status"] == "AUTO_APPROVED" else "❌" if "DENIED" in r["Overall Status"] else "⚠️"
        print(
            f"  {status_icon}  {r['Enrollee ID'][:28]:<28} {r['Encounter Date']}  "
            f"₦{r['Stated Total (₦)']:>9,.2f} → ₦{r['Approved Total (₦)']:>9,.2f}"
            f"  (save ₦{r['Savings (₦)']:,.2f})"
        )
    print("━"*66 + "\n")

    # ── Excel report ──────────────────────────────────────────────────────────
    if output_path:
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # Sheet 1: Summary
                summary_df.to_excel(writer, sheet_name="Batch Summary", index=False)
                ws = writer.sheets["Batch Summary"]
                _style_worksheet(ws, summary_df,
                    status_col="Overall Status",
                    approve_val="AUTO_APPROVED",
                    deny_val="AUTO_DENIED")

                # Sheet 2: Line Detail
                detail_df.to_excel(writer, sheet_name="Line Detail", index=False)
                ws2 = writer.sheets["Line Detail"]
                _style_worksheet(ws2, detail_df,
                    status_col="Status",
                    approve_val="AUTO_APPROVED",
                    deny_val="AUTO_DENIED")

                # Sheet 3: Savings Highlight
                savings_df = detail_df[detail_df["Status"] == "AUTO_DENIED"].copy()
                if not savings_df.empty:
                    savings_df.to_excel(writer, sheet_name="Denied Lines", index=False)
                    ws3 = writer.sheets["Denied Lines"]
                    _style_worksheet(ws3, savings_df,
                        status_col="Status",
                        approve_val="AUTO_APPROVED",
                        deny_val="AUTO_DENIED")

            logger.info(f"Report saved → {output_path}")
            print(f"  📊 Excel report saved: {output_path}\n")
        except ImportError:
            logger.warning("openpyxl not available — skipping Excel report")
        except Exception as e:
            logger.warning(f"Excel write error: {e}")

    return summary_df


def _style_worksheet(ws, df: pd.DataFrame, status_col: str, approve_val: str, deny_val: str):
    """Apply colour coding and auto-width to an openpyxl worksheet."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        GREEN  = PatternFill("solid", fgColor="C6EFCE")
        RED    = PatternFill("solid", fgColor="FFC7CE")
        HEADER = PatternFill("solid", fgColor="2E4057")

        # Header row
        for cell in ws[1]:
            cell.fill      = HEADER
            cell.font      = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data rows
        if status_col in df.columns:
            col_idx = list(df.columns).index(status_col) + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell_val = str(row[col_idx - 1].value or "")
                if approve_val in cell_val:
                    for c in row:
                        c.fill = GREEN
                elif deny_val in cell_val or "DENIED" in cell_val:
                    for c in row:
                        c.fill = RED

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)
    except Exception:
        pass  # styling is best-effort


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def process_claims(
    claims_path: str,
    api_url: str = "http://localhost:8000",
    provider_id: Optional[str] = None,
    hospital_name: Optional[str] = None,
    output_path: Optional[str] = None,
    db_path: Optional[str] = None,
    dry_run: bool = False,
) -> pd.DataFrame:
    """
    Full pipeline: load → batch → resolve → submit → report.
    Returns summary DataFrame.
    """
    # Load
    df = load_claims(claims_path)

    # Detect hospital name from file if not provided
    if not hospital_name and "provider_name" in df.columns:
        hospital_name = df["provider_name"].dropna().iloc[0] if not df["provider_name"].dropna().empty else None

    # Build batches
    batches = build_batches(df)

    # Resolver
    resolver = CodeResolver(db_path)

    if dry_run:
        logger.info("DRY RUN — batches built but not submitted to API")
        for i, b in enumerate(batches, 1):
            print(f"  Batch {i}: {b['enrollee_id']} | {b['encounter_date']} | {len(b['rows'])} rows")
        return pd.DataFrame()

    # Submit each batch
    results = []
    for i, batch in enumerate(batches, 1):
        print(f"\n[{i}/{len(batches)}] {batch['enrollee_id']} — {batch['encounter_date']}")
        result = submit_batch(batch, resolver, api_url, provider_id, hospital_name)
        results.append(result)

    # Report
    summary = generate_report(results, output_path)
    return summary


# ─────────────────────────────────────────────────────────────────────────────
# CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Clearline Claims Batch Processor — vets a hospital claims file through the PA system",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python -m apis.vetting.claims_processor claims.xlsx --provider-id 118 --output report.xlsx
  python -m apis.vetting.claims_processor claims.csv  --api http://localhost:8000 --dry-run
        """,
    )
    parser.add_argument("claims_file", help="Path to claims Excel (.xlsx) or CSV file")
    parser.add_argument("--api",         default="http://localhost:8000",    help="Vetting API base URL")
    parser.add_argument("--provider-id", default=None,                       help="Provider ID / key (e.g. 118)")
    parser.add_argument("--hospital",    default=None,                       help="Hospital name override")
    parser.add_argument("--output",      default=None,                       help="Output Excel report path")
    parser.add_argument("--db",          default=None,                       help="MotherDuck DB path for code resolution")
    parser.add_argument("--dry-run",     action="store_true",                help="Parse and batch only — do not call API")

    args = parser.parse_args()

    if not os.path.exists(args.claims_file):
        print(f"ERROR: File not found: {args.claims_file}", file=sys.stderr)
        sys.exit(1)

    # Auto-generate output path if not given
    output = args.output
    if not output and not args.dry_run:
        base = os.path.splitext(args.claims_file)[0]
        output = f"{base}_vetting_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    process_claims(
        claims_path=args.claims_file,
        api_url=args.api,
        provider_id=args.provider_id,
        hospital_name=args.hospital,
        output_path=output,
        db_path=args.db,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
